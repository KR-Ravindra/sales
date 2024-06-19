import cProfile
from datetime import date
import streamlit as st
from pycel import ExcelCompiler
from openpyxl import load_workbook
import logging
import os
import sys
import time
import threading
from PIL import Image
from products.dynamo import handle_dynamo_all
from products.zippy6 import handle_zippy6
from products.zippy10 import handle_zippy10
from products.zippy30 import handle_zippy30
from products.zippy40 import handle_zippy40
from products.utils import save_to_file, convert_to_meters, convert_to_feet



def pycel_logging_to_console(enable=True):
    if enable:
        logger = logging.getLogger('pycel')
        logger.setLevel('INFO')

        console = logging.StreamHandler(sys.stdout)
        console.setLevel(logging.INFO)
        logger.addHandler(console)

def handle_cover(excel, wb, unit):
    st.markdown("<h1 style='text-align: center; color: red;'>Product Throughput Calculator</h1>", unsafe_allow_html=True)

def handle_sheet2(excel, wb, unit):
    # Similar code for Sheet2
    pass

def loader():
    with st.spinner(text='Hold on tight! :rocket:'):
       time.sleep(0.5)


users = {
    "admin": {
        "password": "adminpassword",
        "role": "admin"
    },
    "user": {
        "password": "userpassword",
        "role": "user"
    }
}
def main(wb, sheet_names, files, excel):
    
    
    st.sidebar.image("./images/AddverbLogo.png", use_column_width=True)
    st.sidebar.title("Throughput Calculator")

    selected_sheet = st.sidebar.selectbox('Select Product', sheet_names)
    unit = st.sidebar.radio("Select Unit", ('Feet (ft)', 'Meters (m)'))
    selected_file = st.sidebar.selectbox('History', files, key='file_select')

    
    if selected_file:
        with open(os.path.join('./files', selected_file), 'rb') as file:
            file_data = file.read()
        st.sidebar.download_button(
            label="Download selected file",
            data=file_data,
            file_name=selected_file
        )
    if selected_sheet == 'Dynamo all':
        handle_dynamo_all(excel, wb, unit)
    elif selected_sheet == 'COVER':
        handle_cover(excel, wb, unit)
    elif selected_sheet == 'Cycle Time ZIppy 6':
        handle_zippy6(excel, wb, unit)
    elif selected_sheet == 'Cycle Time ZIppy 10':
        handle_zippy10(excel, wb, unit)
    elif selected_sheet == 'Cycle Time ZIppy 10':
        handle_zippy10(excel, wb, unit)
    elif selected_sheet == 'Cycle Time ZIppy 30':
        handle_zippy30(excel, wb, unit)
    elif selected_sheet == 'Cycle Time ZIppy 40':
        handle_zippy40(excel, wb, unit)

if __name__ == '__main__':
    st.set_page_config(page_title="Throughput Calculator", page_icon="images/AddverbLogo.png")
    excel = ExcelCompiler(filename='master.xlsx')
    wb = load_workbook('master.xlsx')
    sheet_names = wb.sheetnames
    files = os.listdir('./files')
    main(wb, sheet_names, files, excel)
    